"""Wrappers importables sobre los scripts de extracción existentes.

Permite llamar Phase 1 y Phase 2 como funciones Python sin correr los scripts
como procesos separados. Los scripts originales quedan sin modificar.
"""
from __future__ import annotations

import os
import sys
import threading
from pathlib import Path
from types import SimpleNamespace
from typing import Callable, Optional

# Asegurar que el raíz del proyecto está en sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Lock para setear DEEPSEEK_API_KEY de forma segura entre threads concurrentes
_api_key_lock = threading.Lock()


def run_phase1(input_path: Path, output_path: Path) -> int:
    """Parsea el xlsx con el parser determinístico. Devuelve cantidad de filas."""
    from scripts.extract_descripcion import process_file

    output_path.parent.mkdir(parents=True, exist_ok=True)
    process_file(input_path, output_path)

    # Contar filas del output (sin contar cabecera)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(output_path, read_only=True)
        ws = wb.active
        return max(0, ws.max_row - 1)
    except Exception:
        return 0


def run_phase2(
    raw_path: Path,
    v1_path: Path,
    out_path: Path,
    api_key: str,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Optional[float]:
    """Normaliza con LLM (DeepSeek). Devuelve costo estimado en USD o None."""
    from scripts.llm import vocab as vocab_mod
    from scripts.llm.cache import Cache
    from scripts.llm.client import Stats

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Vocabulario
    vocab_path = PROJECT_ROOT / "data" / "ejemplo.xlsx"
    v = vocab_mod.load(vocab_path)

    cache = Cache()

    args = SimpleNamespace(
        sample=0,
        all=True,
        input=str(raw_path),
        batch_size=10,
        workers=4,
        model=None,
        dry_run=False,
    )

    # Inyectar API key temporalmente en el entorno (seguro con lock para el uso interno)
    with _api_key_lock:
        prev_key = os.environ.get("DEEPSEEK_API_KEY")
        os.environ["DEEPSEEK_API_KEY"] = api_key
        try:
            _run_phase2_inner(raw_path, v1_path, out_path, v, cache, args, on_progress)
        finally:
            if prev_key is None:
                os.environ.pop("DEEPSEEK_API_KEY", None)
            else:
                os.environ["DEEPSEEK_API_KEY"] = prev_key

    # Leer costo del reporte (columna "metrica" == "costo_estimado_usd")
    try:
        import pandas as pd
        rep = pd.read_excel(out_path, sheet_name="_reporte")
        mask = rep["metrica"].astype(str) == "costo_estimado_usd"
        if mask.any():
            return float(rep.loc[mask, "valor"].iloc[0])
    except Exception:
        pass
    return None


def _run_phase2_inner(raw_path, v1_path, out_path, v, cache, args, on_progress):
    """Llama process_file de extract_llm con un on_batch que reporta progreso."""
    import datetime as _dt

    import openpyxl
    import pandas as pd

    from scripts.llm import report, sampler, validate
    from scripts.llm.cache import text_key
    from scripts.llm.client import DeepSeekClient

    HEADER_ROW = 6

    # Replicar load_v1_with_desc de extract_llm
    df = pd.read_excel(v1_path, sheet_name="estructurado")
    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    j_desc = header.index("Descripcion Comercial")
    j_dua = header.index("DUA / DAM")
    duas, descs = [], []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if all(c is None for c in row):
            continue
        duas.append(row[j_dua])
        descs.append(row[j_desc])
    if len(descs) != len(df):
        raise ValueError(f"Desalineación: v1={len(df)} filas, crudo={len(descs)}")
    df["_desc"] = descs
    df["row_key"] = df.apply(
        lambda r: f"{r['dua_dam']}|{r['vin'] if pd.notna(r.get('vin')) else r.get('chasis')}",
        axis=1,
    )

    sub = df.copy()
    sub["_tkey"] = sub["_desc"].map(text_key)
    pendientes = {}
    for _, r in sub.iterrows():
        k = r["_tkey"]
        if k not in cache and k not in pendientes and isinstance(r["_desc"], str):
            pendientes[k] = r["_desc"]

    total_pending = len(pendientes)
    done_count = [0]

    def on_batch(content, batch, keymap, cache_obj):
        from scripts.llm import validate as val
        parsed = val.parse_json_lenient(content)
        items = val.items_by_index(parsed)
        for i, (_, _desc) in enumerate(batch):
            item = items.get(i)
            if item is not None:
                cache_obj.put(keymap[i], item)
        done_count[0] += len(batch)
        if on_progress and total_pending > 0:
            on_progress(done_count[0], total_pending)

    client = DeepSeekClient(v, batch_size=args.batch_size, workers=args.workers)
    pend = dict(pendientes)
    for paso in range(3):
        if not pend:
            break
        client.batch_size = args.batch_size if paso == 0 else 1
        client.run(list(pend.items()), cache, on_batch=on_batch)
        pend = {k: d for k, d in pendientes.items() if k not in cache}

    # Expandir cache -> registros normalizados
    recs = []
    for _, r in sub.iterrows():
        raw_item = cache.get(r["_tkey"])
        rec = validate.normalize_record(raw_item, v) if raw_item else validate.empty_record()
        recs.append(rec)
    norm_df = pd.DataFrame(recs, index=sub.index)
    out = pd.concat([sub.drop(columns=["_desc", "_tkey"]), norm_df], axis=1)
    fecha = _dt.date.today().isoformat()
    modelo_usado = os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-flash")
    out["fuente"] = f"LLM:{modelo_usado}@{fecha}"

    rep_df = report.build(out, client.stats)

    revisar = out[
        out["modelo_flag"].isin(["low", "nomatch", "alias"])
        | (~out["marca_in_vocab"])
        | (~out["traccion_valido"])
        | (~out["combustible_valido"])
        | (~out["clasificacion_valido"])
        | (~out["caja_valido"])
    ]
    nuevos = out[(~out["marca_in_vocab"]) & out["marca_norm"].notna()]
    vocab_nuevo = (
        nuevos.groupby("marca_norm")
        .agg(
            unidades=("marca_norm", "size"),
            sugerencia=("marca_sugerencia", "first"),
            modelos=("modelo_raw_llm", lambda s: ", ".join(
                sorted({str(x) for x in s if pd.notna(x)})[:10]
            )),
        )
        .sort_values("unidades", ascending=False)
        .reset_index()
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name="normalizado_llm", index=False)
        revisar.to_excel(xw, sheet_name="_revisar_llm", index=False)
        vocab_nuevo.to_excel(xw, sheet_name="_vocab_nuevo", index=False)
        rep_df.to_excel(xw, sheet_name="_reporte", index=False)
