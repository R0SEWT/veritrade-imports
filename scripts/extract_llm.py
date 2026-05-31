#!/usr/bin/env python3
"""Fase B — extracción estructurada con LLM (DeepSeek), híbrida y normalizada.

Normaliza 6 campos (marca, modelo, tracción, combustible, clasificación, caja)
contra el vocabulario controlado de ejemplo.xlsx, sobre una muestra de la tabla
determinística v1. El resto de columnas (vin/chasis/números) se conserva de v1.

Uso:
  DEEPSEEK_API_KEY=... python3 scripts/extract_llm.py --sample 300

La API key se lee solo del entorno; nunca se hardcodea. Output marcado como
generado por IA (columna `fuente`); revisar `_revisar_llm` antes de usar.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os
import sys
from pathlib import Path


def load_dotenv(path=".env") -> None:
    """Carga variables de un .env simple (KEY=VALUE) al entorno, sin pisar las ya definidas."""
    p = Path(path)
    if not p.exists():
        return
    for line in p.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, val = line.partition("=")
        os.environ.setdefault(k.strip(), val.strip().strip('"').strip("'"))

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.llm import report, sampler, validate, vocab as vocab_mod  # noqa: E402
from scripts.llm.cache import Cache, text_key  # noqa: E402

INPUTS_DIR = "inputs"    # exports crudos de Veritrade (vehículos)
OUTPUTS_DIR = "outputs"  # <stem>_estructurado.xlsx (entrada v1) y <stem>_normalizado.xlsx (salida)
HEADER_ROW = 6

NORM_COLS = [
    "marca_raw_llm", "marca_norm", "marca_in_vocab", "marca_sugerencia",
    "modelo_raw_llm", "modelo_match", "modelo_score", "modelo_flag",
    "traccion_norm", "traccion_valido", "combustible_norm", "combustible_valido",
    "clasificacion_norm", "clasificacion_valido", "caja_norm", "caja_valido",
]


def load_v1_with_desc(raw_path: Path, v1_path: Path) -> pd.DataFrame:
    """Tabla v1 + texto crudo de la descripción (alineado por orden de fila)."""
    df = pd.read_excel(v1_path, sheet_name="estructurado")

    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    j_dua = header.index("DUA / DAM")
    j_desc = header.index("Descripcion Comercial")
    duas, descs = [], []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if all(c is None for c in row):
            continue
        duas.append(row[j_dua])
        descs.append(row[j_desc])

    if len(descs) != len(df):
        raise SystemExit(f"Desalineación: v1={len(df)} filas, crudo={len(descs)}")
    # sanity: el dua_dam debe coincidir posicionalmente
    mismatch = sum(1 for a, b in zip(df["dua_dam"], duas) if str(a) != str(b))
    if mismatch:
        print(f"Advertencia: {mismatch} dua_dam no coinciden por posición", file=sys.stderr)
    df["_desc"] = descs
    df["row_key"] = df.apply(
        lambda r: f"{r['dua_dam']}|{r['vin'] if pd.notna(r['vin']) else r['chasis']}", axis=1)
    return df


def make_on_batch():
    def on_batch(content, batch, keymap, cache: Cache):
        parsed = validate.parse_json_lenient(content)
        items = validate.items_by_index(parsed)
        for i, (_, desc) in enumerate(batch):
            item = items.get(i)
            if item is not None:
                cache.put(keymap[i], item)
    return on_batch


def process_file(raw_path: Path, v1_path: Path, out_path: Path, v, cache: Cache, args) -> bool:
    print(f"\n=== {raw_path.name} ===")
    df = load_v1_with_desc(raw_path, v1_path)
    # En batch (varios archivos) se procesa completo; --sample solo aplica con un único --input.
    use_sample = bool(args.sample) and not args.all and args.input
    sub = sampler.sample(df, v, n=args.sample) if use_sample else df
    print(f"Filas a procesar: {len(sub)}  (de {len(df)})")
    if "estrato" in sub:
        print("Estratos:", sub["estrato"].value_counts().to_dict())
    # textos únicos no cacheados
    sub = sub.copy()
    sub["_tkey"] = sub["_desc"].map(text_key)
    pendientes = {}
    for _, r in sub.iterrows():
        k = r["_tkey"]
        if k not in cache and k not in pendientes and isinstance(r["_desc"], str):
            pendientes[k] = r["_desc"]
    print(f"Textos únicos pendientes (no cacheados): {len(pendientes)}")

    stats = report.Stats() if hasattr(report, "Stats") else None
    if args.dry_run:
        from scripts.llm.client import Stats
        stats = Stats()
    else:
        from scripts.llm.client import DeepSeekClient
        model = args.model or None
        client = DeepSeekClient(v, **({"model": model} if model else {}),
                                batch_size=args.batch_size, workers=args.workers)
        on_batch = make_on_batch()
        # Pase inicial + reintentos para índices que el LLM haya omitido en su batch.
        pend = dict(pendientes)
        for paso in range(3):
            if not pend:
                break
            client.batch_size = args.batch_size if paso == 0 else 1  # 1 por request al reintentar
            client.run(list(pend.items()), cache, on_batch=on_batch)
            pend = {k: d for k, d in pendientes.items() if k not in cache}
            if paso and pend:
                print(f"  reintento {paso}: quedan {len(pend)} sin respuesta", file=sys.stderr)
        stats = client.stats

    # Expandir cache -> registros normalizados
    recs = []
    for _, r in sub.iterrows():
        raw = cache.get(r["_tkey"])
        rec = validate.normalize_record(raw, v) if raw else validate.empty_record()
        recs.append(rec)
    norm_df = pd.DataFrame(recs, index=sub.index)
    out = pd.concat([sub.drop(columns=["_desc", "_tkey"]), norm_df], axis=1)
    fecha = _dt.date.today().isoformat()
    modelo_usado = args.model or __import__("os").environ.get("DEEPSEEK_MODEL", "deepseek-v4-flash")
    out["fuente"] = f"LLM:{modelo_usado}@{fecha}"

    rep = report.build(out, stats)
    print("\n=== REPORTE ===")
    print(rep.to_string(index=False))

    # Escribir xlsx
    revisar = out[
        out["modelo_flag"].isin(["low", "nomatch", "alias"]) |
        (~out["marca_in_vocab"]) |
        (~out["traccion_valido"]) | (~out["combustible_valido"]) |
        (~out["clasificacion_valido"]) | (~out["caja_valido"])
    ]
    # Vocabulario nuevo: marcas leídas que NO están en ejemplo.xlsx (candidatas a ampliar)
    nuevos = out[(~out["marca_in_vocab"]) & out["marca_norm"].notna()]
    vocab_nuevo = (nuevos.groupby("marca_norm")
                   .agg(unidades=("marca_norm", "size"),
                        sugerencia=("marca_sugerencia", "first"),
                        modelos=("modelo_raw_llm", lambda s: ", ".join(
                            sorted({str(x) for x in s if pd.notna(x)})[:10])))
                   .sort_values("unidades", ascending=False).reset_index())

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name="normalizado_llm", index=False)
        revisar.to_excel(xw, sheet_name="_revisar_llm", index=False)
        vocab_nuevo.to_excel(xw, sheet_name="_vocab_nuevo", index=False)
        rep.to_excel(xw, sheet_name="_reporte", index=False)
    print(f"Escrito: {out_path}  ({len(out)} filas, {len(revisar)} a revisar, "
          f"{len(vocab_nuevo)} marcas nuevas fuera del ejemplo)")
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description="Normalización con LLM (DeepSeek) contra vocabulario (batch).")
    ap.add_argument("--inputs-dir", default=INPUTS_DIR, help="carpeta con exports crudos (default: inputs/)")
    ap.add_argument("--outputs-dir", default=OUTPUTS_DIR, help="carpeta de salida (default: outputs/)")
    ap.add_argument("--input", help="procesar un solo crudo (override de --inputs-dir)")
    ap.add_argument("--vocab", default=None, help="ruta del vocabulario (default: data/ejemplo.xlsx)")
    ap.add_argument("--sample", type=int, default=0, help="muestrear N filas (solo con --input, para pruebas)")
    ap.add_argument("--all", action="store_true", help="forzar procesar todo (default en batch)")
    ap.add_argument("--batch-size", type=int, default=10)
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--model", default=None)
    ap.add_argument("--dry-run", action="store_true", help="no llamar a la API (usa solo cache)")
    args = ap.parse_args()

    load_dotenv()
    v = vocab_mod.load(args.vocab) if args.vocab else vocab_mod.load()

    if args.input:
        srcs = [Path(args.input)]
    else:
        srcs = sorted(Path(args.inputs_dir).glob("*.xlsx"))
    if not srcs:
        print(f"No se encontraron .xlsx en {args.inputs_dir}/", file=sys.stderr)
        return 1

    out_dir = Path(args.outputs_dir)
    cache = Cache()  # compartida entre archivos (clave = texto de descripción)
    ok = 0
    for raw in srcs:
        v1 = out_dir / f"{raw.stem}_estructurado.xlsx"
        if not v1.exists():
            print(f"Falta el estructurado de {raw.name} ({v1}); corre extract_descripcion.py primero.",
                  file=sys.stderr)
            continue
        out = out_dir / f"{raw.stem}_normalizado.xlsx"
        if process_file(raw, v1, out, v, cache, args):
            ok += 1
    print(f"\nListo: {ok}/{len(srcs)} archivos normalizados.")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
